#!/usr/bin/env python3
"""Check actual headers and detection logic"""
from openpyxl import load_workbook
from pathlib import Path

file_path = Path('demodata/11.XLSX')
wb = load_workbook(file_path, data_only=True)
ws = wb.active

headers = []
for col in range(1, min(ws.max_column or 30, 30) + 1):
    cell_value = ws.cell(1, col).value
    if cell_value:
        headers.append(str(cell_value).lower().strip())
    else:
        break  # Stop at empty column

wb.close()

print("File: 11.XLSX")
print(f"Headers: {headers}")
print(f"\nHeaders string: {'|'.join(headers)}")

headers_str = '|'.join(headers)

# Check detection logic from backend
print("\n=== Backend Detection Logic ===")
checks = [
    ('order' in headers_str and 'batch' in headers_str and 'material number' in headers_str, "COOISPI"),
    ('material document' in headers_str and 'material' in headers_str, "MB51"),
    ('purch. order' in headers_str or 'purch order' in headers_str, "ZRMM024"),
    ('billing doc' in headers_str or 'billing document' in headers_str, "ZRSD002"),
    ('delivery' in headers_str and 'sold-to party' in headers_str, "ZRSD004"),
    (('material' in headers_str or 'material code' in headers_str) and 'distribution channel' in headers_str and ('ph 1' in headers_str or 'ph 2' in headers_str or 'ph1' in headers_str or 'ph2' in headers_str), "ZRSD006"),
    ('customer name' in headers_str and 'total target' in headers_str and 'total realization' in headers_str, "ZRFI005"),
    ('salesman name' in headers_str and 'semester' in headers_str and 'target' in headers_str, "TARGET"),
]

for check, file_type in checks:
    print(f"{file_type}: {check}")
    
# Debug ZRSD006 check  
print("\n=== ZRSD006 Check Details ===")
print(f"'material' in headers_str: {'material' in headers_str}")
print(f"'material code' in headers_str: {'material code' in headers_str}")
print(f"'distribution channel' in headers_str: {'distribution channel' in headers_str}")
print(f"'ph 1' in headers_str: {'ph 1' in headers_str}")
print(f"'ph 2' in headers_str: {'ph 2' in headers_str}")
print(f"'ph1' in headers_str: {'ph1' in headers_str}")
print(f"'ph2' in headers_str: {'ph2' in headers_str}")
