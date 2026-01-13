#!/usr/bin/env python3
"""Test openpyxl reading of 11.XLSX"""

import sys
from pathlib import Path
from openpyxl import load_workbook

# File từ upload hay demo?
file_path = Path("uploaded_files/11.xlsx")
print(f"Test File: {file_path}")
print(f"File exists: {file_path.exists()}")

if not file_path.exists():
    # Try alternative path
    file_path = Path("excel_files/zrsd006/11.xlsx")
    print(f"Trying: {file_path}, exists: {file_path.exists()}")

if not file_path.exists():
    print("❌ File not found!")
    sys.exit(1)

try:
    wb = load_workbook(file_path, data_only=True)
    print(f"✓ Workbook opened")
    
    ws = wb.active
    print(f"✓ Worksheet selected: {ws.title}")
    print(f"✓ Max row: {ws.max_row}")
    print(f"✓ Max column: {ws.max_column}")
    
    # Get headers from row 1
    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        header_val = ws.cell(1, col_idx).value
        headers[col_idx] = str(header_val).strip() if header_val else ''
    
    print(f"\n✓ Found {len(headers)} headers:")
    for col_idx, h in headers.items():
        print(f"  Col {col_idx}: {h}")
    
    # Check first few data rows
    print(f"\n✓ First 5 data rows (rows 2-6):")
    for row_idx in range(2, min(7, ws.max_row + 1)):
        row_data = {}
        for col_idx, header in headers.items():
            val = ws.cell(row_idx, col_idx).value
            row_data[header] = str(val).strip() if val else ''
        print(f"  Row {row_idx}: {row_data}")
    
    print(f"\n✓ Total data rows: {ws.max_row - 1}")
    wb.close()
    
except Exception as e:
    print(f"❌ Error: {e}")
    import traceback
    traceback.print_exc()
