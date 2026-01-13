#!/usr/bin/env python3
"""Check actual Excel file structure"""
from openpyxl import load_workbook
from pathlib import Path

file_path = Path('demodata/11.XLSX')
wb = load_workbook(file_path)
ws = wb.active

print(f"File: {file_path.name}")
print(f"Max row: {ws.max_row}")
print(f"Max column: {ws.max_column}")

print("\nFirst 3 rows:")
for row_idx in range(1, min(4, ws.max_row + 1)):
    cells = []
    for col_idx in range(1, min(11, ws.max_column + 1)):
        cell = ws.cell(row_idx, col_idx)
        cells.append(f"{cell.value}")
    print(f"  Row {row_idx}: {cells}")

wb.close()
