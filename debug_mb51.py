#!/usr/bin/env python3
"""Debug MB51 vs ZRMM024 detection issue"""
import openpyxl

# MB51 headers
mb51_file = 'demodata/mb51.XLSX'
wb = openpyxl.load_workbook(mb51_file)
ws = wb.active
mb51_headers = [str(cell.value) for cell in ws[1] if cell.value]
print("MB51 Headers:", mb51_headers[:10])

# ZRMM024 headers
zrmm024_file = 'demodata/zrmm024.XLSX'
wb = openpyxl.load_workbook(zrmm024_file)
ws = wb.active
zrmm024_headers = [str(cell.value) for cell in ws[1] if cell.value]
print("\nZRMM024 Headers:", zrmm024_headers[:10])

# Check MB51 signature from fileDetection.ts
mb51_signature = ['Posting Date', 'Movement Type', 'Material Document', 'Qty in Un. of Entry', 'Storage Location']
print(f"\nMB51 Signature: {mb51_signature}")
mb51_matches = [sig for sig in mb51_signature if sig in mb51_headers]
print(f"MB51 Matches in mb51.XLSX: {len(mb51_matches)}/{len(mb51_signature)} = {mb51_matches}")

# Check ZRMM024 signature
zrmm024_signature = ['Material', 'Material Description', 'Purch. Order']
print(f"\nZRMM024 Signature: {zrmm024_signature}")
mb51_zrmm024_matches = [sig for sig in zrmm024_signature if sig in mb51_headers]
print(f"ZRMM024 Matches in mb51.XLSX: {len(mb51_zrmm024_matches)}/{len(zrmm024_signature)} = {mb51_zrmm024_matches}")
print(f"Match rate: {len(mb51_zrmm024_matches)/len(zrmm024_signature)*100:.0f}% (threshold 60%)")

# Check if Material and Material Description are in MB51
print(f"\n'Material' in mb51.XLSX: {'Material' in mb51_headers}")
print(f"'Material Description' in mb51.XLSX: {'Material Description' in mb51_headers}")
print(f"'Purch. Order' in mb51.XLSX: {'Purch. Order' in mb51_headers}")
