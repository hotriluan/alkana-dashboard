#!/usr/bin/env python3
"""Check with openpyxl directly"""
from openpyxl import load_workbook
from pathlib import Path

file_path = Path('demodata/11.XLSX')
wb = load_workbook(file_path)
ws = wb.active

print(f"Worksheet name: {ws.title}")
print(f"Dimensions: {ws.dimensions}")

# Get headers from row 1
headers = []
for col in range(1, 11):
    cell = ws.cell(1, col)
    headers.append(cell.value)

print(f"\nRow 1 (headers): {headers}")

# Get data from row 2
row2 = []
for col in range(1, 11):
    cell = ws.cell(2, col)
    row2.append(cell.value)

print(f"Row 2 (first data): {row2}")

# Now try with pandas using openpyxl
import pandas as pd
print("\n\n=== Pandas with manual DataFrame ===")

# Read with openpyxl manually
data_rows = []
for row_idx in range(2, min(10, ws.max_row + 1)):
    row_data = []
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row_idx, col_idx)
        row_data.append(cell.value)
    data_rows.append(row_data)

# Create DataFrame with proper headers
header_list = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
df = pd.DataFrame(data_rows, columns=header_list)

print(f"DataFrame shape: {df.shape}")
print(f"DataFrame columns: {list(df.columns)[:5]}")
print(f"First row: {df.iloc[0, :3].tolist()}")

wb.close()
