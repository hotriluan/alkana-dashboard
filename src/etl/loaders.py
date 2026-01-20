"""
Excel Loaders - Extract ALL data from Excel files

CRITICAL REQUIREMENT: Load ALL rows, ALL columns - NO FILTERING
- Store entire row as JSON in raw_data field for safety
- Include source_row for traceability
- No data loss at extraction stage
"""
import pandas as pd
import json
from pathlib import Path
from typing import Dict, Any, List, Optional
from datetime import datetime, date, timedelta
import hashlib

from openpyxl import load_workbook

from sqlalchemy.orm import Session
from sqlalchemy.dialects.postgresql import insert
from sqlalchemy import text

from src.config import EXCEL_FILES
from src.db.models import (
    RawCooispi, RawMb51, RawZrmm024, RawZrsd002,
    RawZrsd004, RawZrsd006, RawZrfi005, RawTarget,
    RawZrpp062, FactProductionPerformanceV2
)


def compute_row_hash(row_dict: Dict) -> str:
    """Compute MD5 hash of row for change detection"""
    # Convert to sorted JSON string for consistent hashing
    json_str = json.dumps(row_dict, sort_keys=True, default=str)
    return hashlib.md5(json_str.encode()).hexdigest()


def safe_str(val) -> Optional[str]:
    """Safely convert value to string, handling NaN"""
    if pd.isna(val):
        return None
    return str(val).strip() if val else None


def safe_int(val) -> Optional[int]:
    """Safely convert value to int, handling NaN"""
    if pd.isna(val):
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def safe_float(val) -> Optional[float]:
    """
    Safely convert value to float, handling NaN and European decimal format.
    Handles both dot (0.882) and comma (0,882) as decimal separator.
    """
    if pd.isna(val):
        return None
    try:
        # Handle European decimal format (comma as separator)
        if isinstance(val, str):
            val = val.replace(',', '.')
        return float(val)
    except (ValueError, TypeError):
        return None


def safe_datetime(val) -> Optional[datetime]:
    """Safely convert value to datetime"""
    if pd.isna(val):
        return None
    try:
        if isinstance(val, datetime):
            return val
        return pd.to_datetime(val)
    except:
        return None


def row_to_json(row: pd.Series) -> Dict:
    """Convert pandas row to JSON-serializable dict"""
    result = {}
    for col in row.index:
        val = row[col]
        if pd.isna(val):
            result[str(col)] = None
        elif isinstance(val, datetime):
            result[str(col)] = val.isoformat()
        else:
            result[str(col)] = val
    return result


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    # remove common punctuation/spaces for robust matching
    for ch in [" ", "\t", "\n", ".", ":", "/", "-", "(", ")", "+", "%"]:
        text = text.replace(ch, "")
    return text


def detect_zrmm024_header_row(file_path: Path, max_scan_rows: int = 10) -> Optional[int]:
    """Detect ZRMM024 header row using openpyxl (1-based Excel row index)."""
    wb = load_workbook(file_path, data_only=True)  # Don't use read_only for header detection
    try:
        ws = wb.active
        expected = {"purchorder", "item", "purchdate"}

        scan_limit = min(max_scan_rows, ws.max_row or max_scan_rows)
        for row_idx in range(1, scan_limit + 1):
            values = [ws.cell(row=row_idx, column=c).value for c in range(1, (ws.max_column or 0) + 1)]
            normalized = {_normalize_header(v) for v in values if v is not None and str(v).strip() != ""}
            if expected.issubset(normalized):
                return row_idx
        return None
    finally:
        wb.close()


class BaseLoader:
    """Base class for Excel loaders"""
    
    def __init__(self, db: Session, mode: str = 'upsert', file_path: Optional[Path] = None):
        """
        Initialize loader
        
        Args:
            db: Database session
            mode: Load mode ('insert' or 'upsert')
                - insert: Load all rows (default behavior)
                - upsert: Update existing, insert new, skip unchanged
            file_path: Path to Excel file (optional, falls back to EXCEL_FILES config)
        """
        self.db = db
        self.mode = mode
        self.file_path = file_path
        self.loaded_count = 0
        self.updated_count = 0
        self.skipped_count = 0
        self.error_count = 0
        self.errors = []
    
    def load(self, **kwargs) -> Dict[str, int]:
        """Load data and return stats"""
        raise NotImplementedError
    
    def get_stats(self) -> Dict[str, Any]:
        return {
            'loaded': self.loaded_count,
            'updated': self.updated_count,
            'skipped': self.skipped_count,
            'errors': self.errors  # Return error list, not count
        }
    
    def upsert_record(self, model_class, business_keys: Dict, record_data: Dict, row_hash: str):
        """
        Upsert a single record based on business keys
        
        Args:
            model_class: SQLAlchemy model (e.g., RawZrsd002)
            business_keys: Dict of business key columns (e.g., {'billing_doc': '12345'})
            record_data: Full record data to insert/update
            row_hash: MD5 hash of record for change detection
        
        Returns:
            'inserted', 'updated', or 'skipped'
        """
        # First check if row_hash already exists (to avoid unique constraint violation)
        existing_by_hash = self.db.query(model_class).filter_by(row_hash=row_hash).first()
        
        if existing_by_hash:
            # Exact same data already exists - skip
            self.skipped_count += 1
            return 'skipped'
        
        # Find existing record by business keys
        existing = self.db.query(model_class).filter_by(**business_keys).first()
        
        if existing:
            # Record exists with different data - update all fields
            for key, value in record_data.items():
                setattr(existing, key, value)
            self.updated_count += 1
            return 'updated'
        else:
            # New record - insert
            new_record = model_class(**record_data)
            self.db.add(new_record)
            self.loaded_count += 1
            return 'inserted'

    def truncate(self):
        """Truncate the associated raw table"""
        # Mapping loader to RawTable model
        # This relies on the convention that Loader handles specific RawTable
        model_map = {
            CooispiLoader: RawCooispi,
            Mb51Loader: RawMb51,
            Zrmm024Loader: RawZrmm024,
            Zrsd002Loader: RawZrsd002,
            Zrsd004Loader: RawZrsd004,
            Zrsd006Loader: RawZrsd006,
            Zrfi005Loader: RawZrfi005,
            TargetLoader: RawTarget
        }
        
        model = model_map.get(type(self))
        if model:
            print(f"  ✨ Clearing {model.__tablename__}...")
            self.db.query(model).delete()
            self.db.commit()



class CooispiLoader(BaseLoader):
    """Load Production Orders from cooispi.XLSX - ALL COLUMNS"""
    
    def load(self) -> Dict[str, int]:
        file_path = self.file_path or EXCEL_FILES['cooispi']
        print(f"Loading {file_path}...")
        
        # Read header manually with openpyxl (pandas sometimes fails to read it)
        try:
            wb = load_workbook(file_path, data_only=True)  # Don't use read_only for header reading
            ws = wb.active
            headers = [ws.cell(1, col).value for col in range(1, (ws.max_column or 18) + 1)]
            wb.close()
            
            # Read data with header=None, then assign header manually
            df = pd.read_excel(file_path, header=None, skiprows=1, dtype=str)
            df.columns = headers[:len(df.columns)]  # Trim headers to match actual columns
            print(f"  Found {len(df)} rows, {len(df.columns)} columns")
            print(f"  Columns: {list(df.columns)}")
        except Exception as e:
            # Fallback to standard read
            print(f"  ⚠ Could not read header manually: {e}")
            df = pd.read_excel(file_path, header=0, dtype=str)
            print(f"  Found {len(df)} rows, {len(df.columns)} columns")
            print(f"  Columns: {list(df.columns)}")
        
        records = []
        for idx, row in df.iterrows():
            try:
                raw_data = row_to_json(row)
                
                # Business key for COOISPI: order (production order number)
                order = safe_str(row.get('Order'))
                
                record_data = {
                    'plant': safe_int(row.get('Plant')),
                    'sales_order': safe_str(row.get('Sales Order')),
                    'order': order,
                    'order_type': safe_str(row.get('Order Type')),
                    'material_number': safe_str(row.get('Material Number')),
                    'release_date_actual': safe_datetime(row.get('Release date (actual)')),
                    'actual_finish_date': safe_datetime(row.get('Actual finish date')),
                    'material_description': safe_str(row.get('Material description')),
                    'bom_alternative': safe_int(row.get('BOM alternative')),
                    'batch': safe_str(row.get('Batch')),
                    'system_status': safe_str(row.get('System Status')),
                    'mrp_controller': safe_str(row.get('MRP controller')),
                    'order_quantity': safe_float(row.get('Order quantity (GMEIN)')),
                    'delivered_quantity': safe_float(row.get('Delivered quantity (GMEIN)')),
                    'unit_of_measure': safe_str(row.get('Unit of measure')),
                    'source_file': str(file_path.name),
                    'source_row': idx + 2,  # +2 for header and 0-index
                    'raw_data': raw_data,
                    'row_hash': compute_row_hash({**raw_data, 'source_file': str(file_path.name)})
                }
                
                if self.mode == 'upsert':
                    # Upsert mode: update existing, insert new, skip unchanged
                    self.upsert_record(
                        RawCooispi,
                        {'order': order},  # Business key
                        record_data,
                        record_data['row_hash']
                    )
                else:
                    # Insert mode: bulk insert all
                    record = RawCooispi(**record_data)
                    records.append(record)
                    self.loaded_count += 1
                
            except Exception as e:
                self.error_count += 1
                self.errors.append(f"Row {idx}: {str(e)}")
        
        # Bulk insert for insert mode
        if self.mode == 'insert' and records:
            self.db.bulk_save_objects(records)
        
        # Commit at the end
        self.db.commit()
        
        print(f"  ✓ Loaded {self.loaded_count} rows, Updated {self.updated_count}, Skipped {self.skipped_count}")
        return self.get_stats()


class Mb51Loader(BaseLoader):
    """Load Material Movements from mb51.XLSX - WITH HEADER (merged cells)"""
    
    def load(self) -> Dict[str, int]:
        file_path = EXCEL_FILES['mb51']
        print(f"Loading {file_path}...")
        
        # File has header in row 1 but pandas can't read it (merged cells issue)
        # Skip row 1 and assign column names manually
        df = pd.read_excel(file_path, header=None, skiprows=1, dtype=str)
        
        # Assign proper column names based on actual header
        df.columns = [
            'Posting Date', 'Movement Type', 'Plant', 'Storage Location',
            'Material', 'Material Description', 'Batch', 'Qty in Un. of Entry',
            'Unit of Entry', 'Cost Center', 'G/L Account', 'Material Document',
            'Text', 'Reference', 'Reason for Movement', 'Purchase Order'
        ][:len(df.columns)]  # Handle if fewer columns exist
        
        print(f"  Found {len(df)} rows, {len(df.columns)} columns")
        print(f"  Columns: {list(df.columns)}")
        
        # Track seen hashes in current batch to prevent intra-batch duplicates
        seen_hashes = set()
        records = []
        
        for idx, row in df.iterrows():
            try:
                # Build raw_data with ALL columns
                raw_data = {str(col): (None if pd.isna(v) else v) for col, v in row.items()}
                
                material_doc = safe_str(row.get('Material Document'))
                
                record_data = {
                    'col_0_posting_date': safe_datetime(row.get('Posting Date')),
                    'col_1_mvt_type': safe_int(row.get('Movement Type')),
                    'col_2_plant': safe_int(row.get('Plant')),
                    'col_3_sloc': safe_int(row.get('Storage Location')),
                    'col_4_material': safe_str(row.get('Material')),
                    'col_5_material_desc': safe_str(row.get('Material Description')),
                    'col_6_batch': safe_str(row.get('Batch')),
                    'col_7_qty': safe_float(row.get('Qty in Un. of Entry')),
                    'col_8_uom': safe_str(row.get('Unit of Entry')),
                    'col_9_cost_center': safe_str(row.get('Cost Center')),
                    'col_10_gl_account': safe_str(row.get('G/L Account')),
                    'col_11_material_doc': material_doc,
                    'col_12_reference': safe_str(row.get('Text')),
                    'col_13_outbound_delivery': safe_str(row.get('Reference')),
                    'col_14': safe_str(row.get('Reason for Movement')),
                    'col_15_purchase_order': safe_str(row.get('Purchase Order')),
                    'source_file': str(file_path.name),
                    'source_row': idx + 2,
                    'raw_data': raw_data,
                    'row_hash': compute_row_hash({**raw_data, 'source_file': str(file_path.name)})
                }
                
                row_hash = record_data['row_hash']
                
                # Skip if already seen in this batch
                if row_hash in seen_hashes:
                    self.skipped_count += 1
                    continue
                
                if self.mode == 'upsert':
                    result = self.upsert_record(RawMb51, {'col_11_material_doc': material_doc}, record_data, row_hash)
                    if result in ('inserted', 'updated'):
                        seen_hashes.add(row_hash)
                else:
                    record = RawMb51(**record_data)
                    records.append(record)
                    seen_hashes.add(row_hash)
                    self.loaded_count += 1
                
            except Exception as e:
                self.error_count += 1
                self.errors.append(f"Row {idx}: {str(e)}")
        
        if self.mode == 'insert' and records:
            self.db.bulk_save_objects(records)
        self.db.commit()
        
        print(f"  ✓ Loaded {self.loaded_count} rows, Updated {self.updated_count}, Skipped {self.skipped_count}")
        return self.get_stats()


class Zrmm024Loader(BaseLoader):
    """Load Purchase Orders from zrmm024.XLSX - ALL 58 COLUMNS"""
    
    def load(self) -> Dict[str, int]:
        file_path = EXCEL_FILES['zrmm024']
        print(f"Loading {file_path}...")

        header_row = detect_zrmm024_header_row(file_path)
        if header_row is None:
            # Fallback: preserve old behavior if header can't be confidently detected
            df = pd.read_excel(file_path, header=None, dtype=str)
            data_start_row = 1
            print("  ⚠ Could not detect header row via openpyxl; using header=None fallback")
        else:
            # Read using the real header row so we don't ingest header as data
            df = pd.read_excel(file_path, header=header_row - 1, dtype=str)
            data_start_row = header_row + 1  # first data row in Excel

        print(f"  Found {len(df)} rows, {len(df.columns)} columns")

        # Build normalized header lookup (first occurrence wins)
        col_lookup: Dict[str, str] = {}
        for col in df.columns:
            key = _normalize_header(col)
            if key and key not in col_lookup:
                col_lookup[key] = str(col)

        def get_by_norm(row: pd.Series, *norm_keys: str):
            for nk in norm_keys:
                col_name = col_lookup.get(nk)
                if col_name is not None:
                    return row.get(col_name)
            return None
        
        records = []
        for idx, row in df.iterrows():
            try:
                # Store ALL columns in raw_data JSON (by header name when available)
                raw_data = row_to_json(row)

                purch_order = safe_str(
                    get_by_norm(row, "purchorder")
                    if header_row is not None
                    else (row.iloc[0] if len(row) > 0 else None)
                )
                item = safe_int(
                    get_by_norm(row, "item")
                    if header_row is not None
                    else (row.iloc[1] if len(row) > 1 else None)
                )

                record_data = {
                    'purch_order': purch_order,
                    'item': item,
                    'purch_date': safe_datetime(
                        get_by_norm(row, "purchdate")
                        if header_row is not None
                        else (row.iloc[2] if len(row) > 2 else None)
                    ),
                    'suppl_plant': safe_int(get_by_norm(row, "supplplant")) if header_row is not None else None,
                    'dest_plant': safe_int(get_by_norm(row, "destplant")) if header_row is not None else None,
                    'material': safe_str(get_by_norm(row, "material")) if header_row is not None else None,
                    'material_desc': safe_str(get_by_norm(row, "materialdescription")) if header_row is not None else None,
                    'qty_order': safe_float(get_by_norm(row, "qtyorder")) if header_row is not None else None,
                    'gross_weight': safe_float(get_by_norm(row, "grossweight")) if header_row is not None else None,
                    'tonnage_order': safe_float(get_by_norm(row, "tonnageorder")) if header_row is not None else None,
                    'qty_order_tol': safe_float(get_by_norm(row, "qtyordertol")) if header_row is not None else None,
                    'delivery_date': safe_datetime(get_by_norm(row, "deliverydate")) if header_row is not None else None,
                    'qty_gi': safe_float(get_by_norm(row, "qtygi")) if header_row is not None else None,
                    'tonnage_gi': safe_float(get_by_norm(row, "tonnagegi")) if header_row is not None else None,
                    'qty_receipt': safe_float(get_by_norm(row, "qtyreceipt")) if header_row is not None else None,
                    'source_file': str(file_path.name),
                    'source_row': idx + data_start_row,
                    'raw_data': raw_data,
                    'row_hash': compute_row_hash({**raw_data, 'source_file': str(file_path.name)})
                }
                
                if self.mode == 'upsert':
                    self.upsert_record(RawZrmm024, {'purch_order': purch_order, 'item': item}, record_data, record_data['row_hash'])
                else:
                    record = RawZrmm024(**record_data)
                    records.append(record)
                    self.loaded_count += 1
                
            except Exception as e:
                self.error_count += 1
                self.errors.append(f"Row {idx}: {str(e)}")
        
        if self.mode == 'insert' and records:
            self.db.bulk_save_objects(records)
        self.db.commit()
        
        print(f"  ✓ Loaded {self.loaded_count} rows, Updated {self.updated_count}, Skipped {self.skipped_count} (ALL 58 columns in raw_data)")
        return self.get_stats()


class Zrsd002Loader(BaseLoader):
    """Load Billing Documents from zrsd002.XLSX - ALL 30 COLUMNS"""
    
    def load(self) -> Dict[str, int]:
        file_path = EXCEL_FILES['zrsd002']
        print(f"Loading {file_path}...")
        
        # File has hidden/merged row 0 (all NaN), row 1 has headers (openpyxl can see it)
        # Read headers from openpyxl, then load data with pandas
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        wb.close()
        
        # Read data starting from row 2 (skip hidden row 0 and header row 1)
        df = pd.read_excel(file_path, header=None, skiprows=1, dtype=str, names=headers)
        print(f"  Found {len(df)} rows, {len(df.columns)} columns")
        print(f"  Columns: {list(df.columns)}")
        
        records = []
        for idx, row in df.iterrows():
            try:
                raw_data = row_to_json(row)
                
                billing_document = safe_str(row.get('Billing Document'))
                billing_item = safe_int(row.get('Billing Item'))
                
                record_data = {
                    'billing_date': safe_datetime(row.get('Billing Date')),
                    'billing_document': billing_document,
                    'billing_item': billing_item,
                    'sloc': safe_str(row.get('Sloc')),
                    'sales_office': safe_str(row.get('Sales Office')),
                    'dist_channel': safe_str(row.get('Dist Channel')),
                    'customer_name': safe_str(row.get('Name of Bill to')),
                    'cust_group': safe_str(row.get('Cust. Group')),
                    'salesman_name': safe_str(row.get('Salesman Name')),
                    'material': safe_str(row.get('Material')),
                    'material_desc': safe_str(row.get('Description')),
                    'prod_hierarchy': safe_str(row.get('Prod. Hierarchy')),
                    'billing_qty': safe_float(row.get('Billing Qty')),
                    'sales_unit': safe_str(row.get('Sales Unit')),
                    'currency': safe_str(row.get('Curr')),
                    'exchange_rate': safe_float(row.get('Exchange Rate')),
                    'price': safe_float(row.get('Price')),
                    'total_price': safe_float(row.get('Total Price')),
                    'discount_item': safe_float(row.get('Discount Item')),
                    'net_value': safe_float(row.get('Net Value')),
                    'tax': safe_float(row.get('Tax')),
                    'total': safe_float(row.get('Total')),
                    'net_weight': safe_float(row.get('Net Weight')),
                    'weight_unit': safe_str(row.get('Weight Unit')),
                    'volume': safe_float(row.get('Volum')),
                    'volume_unit': safe_str(row.get('Volum Unit')),
                    'so_number': safe_str(row.get('SO No.')),
                    'so_date': safe_datetime(row.get('SO Date.')),
                    'doc_reference_od': safe_str(row.get('Doc Reference (OD).')),
                    'source_file': str(file_path.name),
                    'source_row': idx + 2,
                    'raw_data': raw_data,
                    'row_hash': compute_row_hash(raw_data)  # Don't include source_file - billing_doc+item is unique
                }
                
                if self.mode == 'upsert':
                    self.upsert_record(RawZrsd002, {'billing_document': billing_document, 'billing_item': billing_item}, record_data, record_data['row_hash'])
                else:
                    record = RawZrsd002(**record_data)
                    records.append(record)
                    self.loaded_count += 1
                
            except Exception as e:
                self.error_count += 1
                self.errors.append(f"Row {idx}: {str(e)}")
        
        if self.mode == 'insert' and records:
            self.db.bulk_save_objects(records)
        self.db.commit()
        
        print(f"  ✓ Loaded {self.loaded_count} rows, Updated {self.updated_count}, Skipped {self.skipped_count}")
        return self.get_stats()


class Zrsd004Loader(BaseLoader):
    """Load Delivery Documents from zrsd004.XLSX - ALL 34 COLUMNS"""
    
    def load(self) -> Dict[str, int]:
        file_path = EXCEL_FILES['zrsd004']
        print(f"Loading {file_path}...")
        
        # File has formatted headers in row 1 that pandas cannot parse
        # Skip row 1 and assign column names manually (same pattern as Mb51Loader)
        df = pd.read_excel(file_path, header=None, skiprows=1, dtype=str)
        
        # Assign all 34 column names from actual Excel structure
        df.columns = [
            'Delivery Date', 'Actual GI Date', 'Delivery', 'SO Reference',
            'Req. Type', 'Delivery Type', 'Shipping Point', 'Sloc',
            'Sales Office', 'Dist. Channel', 'Cust. Group', 'Sold-to Party',
            'Ship-to Party', 'Name of Ship-to', 'City of Ship-to',
            'Regional Stru. Grp.', 'Transportation Zone', 'Salesman ID',
            'Salesman Name', 'Material', 'Description', 'Delivery Qty',
            'Tonase', 'Tonase Unit', 'Actual Delivery Qty', 'Sales Unit',
            'Net Weight', 'Weight Unit', 'Volume', 'Volume Unit',
            'Created By', 'Product Hierarchy', 'Line Item',
            'Total Movement Goods Stat'
        ][:len(df.columns)]  # Handle if Excel has fewer columns
        
        print(f"  Found {len(df)} rows, {len(df.columns)} columns")
        print(f"  Columns: {list(df.columns)}")
        
        records = []
        for idx, row in df.iterrows():
            try:
                raw_data = row_to_json(row)
                
                delivery = safe_str(row.get('Delivery'))
                line_item = safe_int(row.get('Line Item'))
                
                record_data = {
                    'delivery_date': safe_datetime(row.get('Delivery Date')),
                    'actual_gi_date': safe_datetime(row.get('Actual GI Date')),
                    'delivery': delivery,
                    'line_item': line_item,
                    'so_reference': safe_str(row.get('SO Reference')),
                    'shipping_point': safe_str(row.get('Shipping Point')),
                    'sloc': safe_str(row.get('Sloc')),  # Fixed: was 'SLoc'
                    'sales_office': safe_str(row.get('Sales Office')),
                    'dist_channel': safe_str(row.get('Dist. Channel')),  # Fixed: was 'Distribution Channel'
                    'cust_group': safe_str(row.get('Cust. Group')),
                    'sold_to_party': safe_str(row.get('Sold-to Party')),
                    'ship_to_party': safe_str(row.get('Ship-to Party')),
                    'ship_to_name': safe_str(row.get('Name of Ship-to')),  # Fixed: was 'Ship-to Name'
                    'ship_to_city': safe_str(row.get('City of Ship-to')),  # Fixed: was 'Ship-to City'
                    'salesman_id': safe_str(row.get('Salesman ID')),
                    'salesman_name': safe_str(row.get('Salesman Name')),
                    'material': safe_str(row.get('Material')),
                    'material_desc': safe_str(row.get('Description')),  # Fixed: was 'Material Description'
                    'delivery_qty': safe_float(row.get('Delivery Qty')),
                    'tonase': safe_float(row.get('Tonase')),
                    'tonase_unit': safe_str(row.get('Tonase Unit')),
                    'net_weight': safe_float(row.get('Net Weight')),
                    'volume': safe_float(row.get('Volume')),
                    'prod_hierarchy': safe_str(row.get('Product Hierarchy')),  # Fixed: was 'Prod. Hierarchy'
                    'source_file': str(file_path.name),
                    'source_row': idx + 2,
                    'raw_data': raw_data,
                    'row_hash': compute_row_hash({**raw_data, 'source_file': str(file_path.name)})
                }
                
                if self.mode == 'upsert':
                    self.upsert_record(RawZrsd004, {'delivery': delivery, 'line_item': line_item}, record_data, record_data['row_hash'])
                else:
                    record = RawZrsd004(**record_data)
                    records.append(record)
                    self.loaded_count += 1
                
            except Exception as e:
                self.error_count += 1
                self.errors.append(f"Row {idx}: {str(e)}")
        
        if self.mode == 'insert' and records:
            self.db.bulk_save_objects(records)
        self.db.commit()
        
        print(f"  ✓ Loaded {self.loaded_count} rows, Updated {self.updated_count}, Skipped {self.skipped_count}")
        return self.get_stats()


class Zrsd006Loader(BaseLoader):
    """Load Material Master from zrsd006.XLSX - ALL COLUMNS"""
    
    def load(self) -> Dict[str, int]:
        # Use self.file_path if provided (from upload), otherwise use default config
        file_path = self.file_path or EXCEL_FILES['zrsd006']
        print(f"Loading {file_path}...")
        print(f"  DEBUG: self.file_path = {self.file_path}")
        print(f"  DEBUG: file_path.exists() = {file_path.exists() if file_path else 'N/A'}")
        
        # Read Excel using openpyxl directly (pandas fails with these files)
        try:
            if not file_path:
                raise ValueError("file_path is None")
            if not file_path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")
                
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # Extract headers from row 1
            headers = {}
            for col_idx in range(1, ws.max_column + 1):
                header_val = ws.cell(1, col_idx).value
                if header_val:
                    headers[col_idx] = str(header_val).strip()
            
            print(f"  Found {len(headers)} header columns")
            
            # Extract data from rows 2 onwards  
            records = []
            for row_idx in range(2, ws.max_row + 1):
                row = {}
                for col_idx, header_name in headers.items():
                    cell_val = ws.cell(row_idx, col_idx).value
                    row[header_name] = str(cell_val).strip() if cell_val else ''
                
                try:
                    raw_data = row
                    
                    # Use actual column names from file
                    material = safe_str(row.get('Material Code'))
                    dist_channel = safe_str(row.get('Distribution Channel'))
                    
                    record_data = {
                        'material': material,
                        'material_desc': safe_str(row.get('Mat. Description')),
                        'dist_channel': dist_channel,
                        'uom': safe_str(row.get('UOM')),
                        'ph1': safe_str(row.get('PH 1')),
                        'ph1_desc': safe_str(row.get('Division')),
                        'ph2': safe_str(row.get('PH 2')),
                        'ph2_desc': safe_str(row.get('Business')),
                        'ph3': safe_str(row.get('PH 3')),
                        'ph3_desc': safe_str(row.get('Sub Business')),
                        'ph4': safe_str(row.get('PH 4')),
                        'ph4_desc': safe_str(row.get('Product Group')),
                        'ph5': safe_str(row.get('PH 5')),
                        'ph5_desc': safe_str(row.get('Product Group 1')),
                        'ph6': safe_str(row.get('PH 6')),
                        'ph6_desc': safe_str(row.get('Product Group 2')),
                        'ph7': safe_str(row.get('PH 7')),
                        'ph7_desc': safe_str(row.get('Series')),
                        'source_file': str(file_path.name),
                        'source_row': row_idx,
                        'raw_data': raw_data,
                        'row_hash': compute_row_hash({**raw_data, 'source_file': str(file_path.name)})
                    }
                    
                    if self.mode == 'upsert':
                        self.upsert_record(RawZrsd006, {'material': material, 'dist_channel': dist_channel}, record_data, record_data['row_hash'])
                    else:
                        record = RawZrsd006(**record_data)
                        records.append(record)
                        self.loaded_count += 1
                    
                except Exception as e:
                    self.error_count += 1
                    self.errors.append(f"Row {row_idx}: {str(e)}")
            
            wb.close()
            
            if self.mode == 'insert' and records:
                self.db.bulk_save_objects(records)
                
        except Exception as e:
            print(f"  ✗ Error reading file: {e}")
            self.error_count += 1
            return self.get_stats()
        self.db.commit()
        
        print(f"  ✓ Loaded {self.loaded_count} rows, Updated {self.updated_count}, Skipped {self.skipped_count}")
        
        # AUTO-POPULATE dim_product_hierarchy (V4: Centralized Smart Ingestion)
        # This enables star schema JOINs for Production Yield V3.5+
        print("\n  🔄 Auto-populating dim_product_hierarchy...")
        try:
            dim_stats = self.load_to_dimension(file_path)
            print(f"  ✓ Dimension table updated: {dim_stats.get('loaded', 0)} materials")
        except Exception as e:
            print(f"  ⚠️  Warning: Failed to update dimension table: {e}")
            # Don't fail the whole load if dimension update fails
        
        return self.get_stats()
    
    def load_to_dimension(self, file_path: Optional[Path] = None) -> Dict[str, Any]:
        """
        Load master data into dim_product_hierarchy dimension table.
        
        Strategy: UPSERT with PostgreSQL ON CONFLICT DO UPDATE
        Key Sanitization: material_code.lstrip('0') to match fact table format
        
        Purpose: Enable star schema JOIN for brand/grade categorization
        """
        from src.db.models import DimProductHierarchy
        
        # Use provided file_path, then self.file_path, then default config
        file_path = file_path or self.file_path or EXCEL_FILES.get('zrsd006')
        if not file_path or not file_path.exists():
            print(f"  ✗ File not found: {file_path}")
            self.error_count += 1
            return self.get_stats()
        
        print(f"Loading {file_path} into dim_product_hierarchy...")
        
        # Read with openpyxl directly to properly handle headers
        # NOTE: pandas.read_excel() fails to read headers from these files
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # Extract headers from row 1
            headers = {}
            for col_idx in range(1, ws.max_column + 1):
                header_val = ws.cell(1, col_idx).value
                if header_val:
                    headers[col_idx] = str(header_val).strip()
            
            # Extract data from rows 2 onwards
            rows_data = []
            for row_idx in range(2, ws.max_row + 1):
                row = {}
                for col_idx, header_name in headers.items():
                    cell_val = ws.cell(row_idx, col_idx).value
                    row[header_name] = str(cell_val).strip() if cell_val else ''
                rows_data.append(row)
            
            wb.close()
            
            # Convert to DataFrame for consistency
            df = pd.DataFrame(rows_data)
        except Exception as e:
            print(f"  ✗ Failed to read Excel: {e}")
            self.error_count += 1
            return self.get_stats()
        
        print(f"  Found {len(df)} rows, {len(df.columns)} columns")
        
        # UPSERT SQL (PostgreSQL native)
        upsert_sql = text("""
            INSERT INTO dim_product_hierarchy (
                material_code, material_description,
                ph_level_1, ph_level_2, ph_level_3,
                created_at, updated_at
            ) VALUES (
                :material_code, :material_description,
                :ph_level_1, :ph_level_2, :ph_level_3,
                NOW(), NOW()
            )
            ON CONFLICT (material_code) DO UPDATE SET
                material_description = EXCLUDED.material_description,
                ph_level_1 = EXCLUDED.ph_level_1,
                ph_level_2 = EXCLUDED.ph_level_2,
                ph_level_3 = EXCLUDED.ph_level_3,
                updated_at = NOW()
        """)
        
        for idx, row in df.iterrows():
            try:
                # Get material code and sanitize (CRITICAL: remove leading zeros)
                material_raw = safe_str(row.get('Material Code'))
                if not material_raw:
                    self.skipped_count += 1
                    continue
                
                # Clean material code (match format in fact table)
                material_code = material_raw.lstrip('0') if material_raw else None
                if not material_code:
                    self.skipped_count += 1
                    continue
                
                # Extract PH levels (use descriptions for better readability)
                params = {
                    'material_code': material_code,
                    'material_description': safe_str(row.get('Mat. Description')),
                    'ph_level_1': safe_str(row.get('Division')),      # PH 1 Desc
                    'ph_level_2': safe_str(row.get('Business')),      # PH 2 Desc
                    'ph_level_3': safe_str(row.get('Sub Business')),  # PH 3 Desc (Brand/Grade)
                }
                
                self.db.execute(upsert_sql, params)
                self.loaded_count += 1
                
            except Exception as e:
                self.error_count += 1
                self.errors.append(f"Row {idx}: {str(e)}")
        
        self.db.commit()
        
        print(f"  ✓ Upserted {self.loaded_count} materials, Skipped {self.skipped_count}, Errors {len(self.errors)}")
        return self.get_stats()


class Zrfi005Loader(BaseLoader):
    """Load AR Aging from ZRFI005.XLSX - ALL 20 COLUMNS"""
    
    def load(self, snapshot_date: Optional[date] = None) -> Dict[str, int]:
        """
        Load AR data with upsert logic
        
        Args:
            snapshot_date: Date of AR snapshot
        
        Business Rule (NEW):
            - KEEP all historical snapshots (never delete)
            - For same snapshot date: upsert (update if exists, insert if new)
            - This allows viewing historical data and comparing snapshots
            
            Example:
                * Upload 01/05 → Insert 129 records (snapshot 01/05)
                * Upload 01/06 → Insert 129 records (snapshot 01/06)
                * Re-upload 01/06 → Update existing records (same snapshot date, same business key)
                * Upload 01/07 → Insert 98 records (snapshot 01/07)
                * Can view AR aging for any of: 01/05, 01/06, or 01/07
        """
        file_path = self.file_path or EXCEL_FILES['zrfi005']
        print(f"Loading {file_path}...")
        print(f"  📅 Snapshot date: {snapshot_date or 'Not specified'}")
        print(f"  📝 Mode: {self.mode} (upsert - keep history)")
        
        df = pd.read_excel(file_path, header=0, dtype=str)
        print(f"  Found {len(df)} rows, {len(df.columns)} columns")
        
        records = []
        for idx, row in df.iterrows():
            try:
                raw_data = row_to_json(row)
                
                # Business key for AR: customer_name + snapshot_date
                customer_name = safe_str(row.get('Customer Name'))
                
                record_data = {
                    'dist_channel': safe_str(row.get('Distribution Channel')),
                    'cust_group': safe_str(row.get('Customer Group')),  # Fixed: was 'Cust. Group'
                    'salesman_name': safe_str(row.get('Salesman Name')),
                    'customer_name': customer_name,
                    'currency': safe_str(row.get('Currency')),  # Fixed: was 'Curr'
                    'target_1_30': safe_float(row.get('Target 1-30 Days')),
                    'target_31_60': safe_float(row.get('Target 31-60 Days')),
                    'target_61_90': safe_float(row.get('Target 61 - 90 Days')),  # Fixed: added spaces
                    'target_91_120': safe_float(row.get('Target 91 - 120 Days')),  # Fixed: added spaces
                    'target_121_180': safe_float(row.get('Target 121 - 180 Days')),  # Fixed: added spaces
                    'target_over_180': safe_float(row.get('Target > 180 Days')),
                    'total_target': safe_float(row.get('Total Target')),
                    'realization_not_due': safe_float(row.get('Realization Not Due')),
                    'realization_1_30': safe_float(row.get('Realization 1 - 30 Days')),  # Fixed: added spaces
                    'realization_31_60': safe_float(row.get('Realization 31 - 60 Days')),  # Fixed: added spaces
                    'realization_61_90': safe_float(row.get('Realization 61 - 90 Days')),  # Fixed: added spaces
                    'realization_91_120': safe_float(row.get('Realization 91 - 120 Days')),  # Fixed: added spaces
                    'realization_121_180': safe_float(row.get('Realization 121 - 180 Days')),  # Fixed: added spaces
                    'realization_over_180': safe_float(row.get('Realization > 180 Days')),
                    'total_realization': safe_float(row.get('Total Realization')),
                    'snapshot_date': snapshot_date,
                    'source_file': str(file_path.name),
                    'source_row': idx + 2,
                    'raw_data': raw_data,
                }
                
                # Compute row_hash ONLY from business data (exclude metadata: source_file, source_row)
                # This ensures same data uploaded twice has same hash
                hash_data = {
                    'customer_name': customer_name,
                    'dist_channel': record_data['dist_channel'],
                    'cust_group': record_data['cust_group'],
                    'salesman_name': record_data['salesman_name'],
                    'total_target': record_data['total_target'],
                    'total_realization': record_data['total_realization'],
                    'snapshot_date': snapshot_date.isoformat() if snapshot_date else None
                }
                record_data['row_hash'] = compute_row_hash(hash_data)
                
                if self.mode == 'upsert':
                    # Upsert mode: update existing record with same business key
                    # Business key: customer + distribution channel + customer group + salesman + snapshot date
                    # (one customer can have multiple records for different channels/groups)
                    self.upsert_record(
                        RawZrfi005,
                        {
                            'customer_name': customer_name,
                            'dist_channel': record_data['dist_channel'],
                            'cust_group': record_data['cust_group'],
                            'salesman_name': record_data['salesman_name'],
                            'snapshot_date': snapshot_date
                        },
                        record_data,
                        record_data['row_hash']
                    )
                else:
                    # Insert mode: bulk insert all
                    record = RawZrfi005(**record_data)
                    records.append(record)
                    self.loaded_count += 1
                
            except Exception as e:
                self.error_count += 1
                self.errors.append(f"Row {idx}: {str(e)}")
        
        # Bulk insert for insert mode
        if self.mode == 'insert' and records:
            self.db.bulk_save_objects(records)
        
        # Commit at the end
        self.db.commit()
        
        print(f"  ✓ Loaded {self.loaded_count} rows, Updated {self.updated_count}, Skipped {self.skipped_count}")
        return self.get_stats()


class TargetLoader(BaseLoader):
    """Load Sales Targets from target.xlsx - ALL 4 COLUMNS"""
    
    def load(self) -> Dict[str, int]:
        file_path = EXCEL_FILES['target']
        print(f"Loading {file_path}...")
        
        df = pd.read_excel(file_path, header=0, dtype=str)
        print(f"  Found {len(df)} rows, {len(df.columns)} columns")
        print(f"  Columns: {list(df.columns)}")
        
        records = []
        for idx, row in df.iterrows():
            try:
                raw_data = row_to_json(row)
                
                salesman_name = safe_str(row.get('Salesman Name'))
                semester = safe_int(row.get('Semester'))
                year = safe_int(row.get('Year'))
                
                record_data = {
                    'salesman_name': salesman_name,
                    'semester': semester,
                    'year': year,
                    'target': safe_float(row.get('Target')),
                    'source_file': str(file_path.name),
                    'source_row': idx + 2,
                    'raw_data': raw_data,
                    'row_hash': compute_row_hash({**raw_data, 'source_file': str(file_path.name)})
                }
                
                if self.mode == 'upsert':
                    self.upsert_record(RawTarget, {'salesman_name': salesman_name, 'semester': semester, 'year': year}, record_data, record_data['row_hash'])
                else:
                    record = RawTarget(**record_data)
                    records.append(record)
                    self.loaded_count += 1
                
            except Exception as e:
                self.error_count += 1
                self.errors.append(f"Row {idx}: {str(e)}")
        
        if self.mode == 'insert' and records:
            self.db.bulk_save_objects(records)
        self.db.commit()
        
        print(f"  ✓ Loaded {self.loaded_count} rows, Updated {self.updated_count}, Skipped {self.skipped_count}")
        return self.get_stats()


class Zrpp062Loader(BaseLoader):
    """
    Load Production Performance Data from zrpp062.XLSX (Variance Analysis V3)
    
    V3 FEATURES:
    - UPSERT using PostgreSQL ON CONFLICT DO UPDATE
    - reference_date parameter for historical tracking
    - Unique constraint: (process_order_id, batch_id)
    
    CRITICAL DATA SANITIZATION:
    - Order SFG Liquid: .lstrip('0') to remove leading zeros (e.g., '0100255' -> '100255')
    - Process Order: Convert float to string (e.g., 100255.0 -> '100255')
    - Numeric fields: Convert empty/NaN to None (not 0)
    
    Loads into both:
    1. raw_zrpp062 (staging, all columns as-is)
    2. fact_production_performance_v2 (analytical, with UPSERT)
    """
    
    def _clean_order_id(self, val) -> Optional[str]:
        """
        Clean order ID: remove leading zeros and handle float conversion
        e.g., '0100255' -> '100255', 100255.0 -> '100255'
        """
        if pd.isna(val):
            return None
        str_val = str(val).strip()
        if '.' in str_val:
            try:
                str_val = str(int(float(str_val)))
            except (ValueError, TypeError):
                pass
        cleaned = str_val.lstrip('0')
        return cleaned if cleaned else None
    
    def _clean_process_order(self, val) -> Optional[str]:
        """
        Clean Process Order: handle float conversion
        e.g., 100255.0 -> '100255'
        """
        if pd.isna(val):
            return None
        str_val = str(val).strip()
        if '.' in str_val:
            try:
                str_val = str(int(float(str_val)))
            except (ValueError, TypeError):
                pass
        return str_val if str_val else None
    
    def load_with_period(self, file_path: Path, reference_date: date) -> Dict[str, int]:
        """
        V3 UPSERT Load: Load data with a specific reporting period (reference_date).
        
        Uses PostgreSQL ON CONFLICT DO UPDATE for efficient upsert.
        
        Args:
            file_path: Path to Excel file
            reference_date: First day of the reporting month (e.g., 2026-01-01 for Jan 2026)
        
        Returns:
            Dict with load statistics {loaded, updated, skipped, errors}
        """
        if not file_path or not file_path.exists():
            print(f"  ✗ File not found: {file_path}")
            return self.get_stats()
        
        print(f"Loading {file_path} for period {reference_date}...")
        
        # Read Excel file with openpyxl engine for better compatibility
        try:
            df = pd.read_excel(file_path, header=0, dtype=str, engine='openpyxl')
        except Exception as e:
            print(f"  ✗ Failed to read Excel file: {e}")
            self.error_count += 1
            return self.get_stats()
        
        print(f"  Found {len(df)} rows, {len(df.columns)} columns")
        print(f"  Column names: {list(df.columns)[:10]}...")  # Show first 10 columns
        
        # Prepare UPSERT SQL for fact table
        upsert_sql = text("""
            INSERT INTO fact_production_performance_v2 (
                process_order_id, batch_id, material_code, material_description,
                parent_order_id, mrp_controller, product_group_1, product_group_2,
                output_actual_kg, input_actual_kg, process_order_qty,
                loss_kg, loss_pct, sg_theoretical, sg_actual,
                variant_prod_sfg_pct, variant_fg_pct, reference_date, created_at, updated_at
            ) VALUES (
                :process_order_id, :batch_id, :material_code, :material_description,
                :parent_order_id, :mrp_controller, :product_group_1, :product_group_2,
                :output_actual_kg, :input_actual_kg, :process_order_qty,
                :loss_kg, :loss_pct, :sg_theoretical, :sg_actual,
                :variant_prod_sfg_pct, :variant_fg_pct, :reference_date, NOW(), NOW()
            )
            ON CONFLICT (process_order_id, batch_id) DO UPDATE SET
                material_code = EXCLUDED.material_code,
                material_description = EXCLUDED.material_description,
                parent_order_id = EXCLUDED.parent_order_id,
                mrp_controller = EXCLUDED.mrp_controller,
                product_group_1 = EXCLUDED.product_group_1,
                product_group_2 = EXCLUDED.product_group_2,
                output_actual_kg = EXCLUDED.output_actual_kg,
                input_actual_kg = EXCLUDED.input_actual_kg,
                process_order_qty = EXCLUDED.process_order_qty,
                loss_kg = EXCLUDED.loss_kg,
                loss_pct = EXCLUDED.loss_pct,
                sg_theoretical = EXCLUDED.sg_theoretical,
                sg_actual = EXCLUDED.sg_actual,
                variant_prod_sfg_pct = EXCLUDED.variant_prod_sfg_pct,
                variant_fg_pct = EXCLUDED.variant_fg_pct,
                reference_date = EXCLUDED.reference_date,
                updated_at = NOW()
        """)
        
        raw_records = []
        
        for idx, row in df.iterrows():
            try:
                raw_data = row_to_json(row)
                
                # === Key Identifiers (with cleaning) ===
                process_order = self._clean_process_order(row.get('Process Order'))
                batch = safe_str(row.get('Batch'))
                material = safe_str(row.get('Material'))
                material_description = safe_str(row.get('Material Description'))
                order_sfg_liquid = self._clean_order_id(row.get('Order SFG Liquid'))
                
                # Skip rows without process_order
                if not process_order:
                    self.skipped_count += 1
                    continue
                
                # === Build raw record ===
                raw_record_data = {
                    'process_order': process_order,
                    'batch': batch,
                    'material': material,
                    'material_description': material_description,
                    'order_sfg_liquid': order_sfg_liquid,
                    'mrp_controller': safe_str(row.get('MRP Controller')),
                    'product_group_1': safe_str(row.get('Product Group 1')),
                    'product_group_2': safe_str(row.get('Product Group 2')),
                    'qty_order_sfg_liquid': safe_float(row.get('Qty Order SFG Liquid')),
                    'process_order_qty': safe_float(row.get('Process Order Qty')),
                    'uom': safe_str(row.get('UoM')),
                    'bom_alt': safe_str(row.get('BOM Alt')),
                    'bom_text': safe_str(row.get('BOM Text')),
                    'group_recipe': safe_str(row.get('Group Recipe')),
                    'gi_packaging_to_order': safe_float(row.get('GI Packaging to Order')),
                    'gi_sfg_liquid_to_order': safe_float(row.get('GI SFG Liquid to Order')),
                    'gr_qty_to_0201': safe_float(row.get('GR Qty to 0201')),
                    'tonase_alkana_0201': safe_float(row.get('Tonase Alkana(0201)')),
                    'gr_by_product': safe_float(row.get('GR by Product')),
                    'sg_theoretical': safe_float(row.get('SG Theoretical')),
                    'sg_actual': safe_float(row.get('SG Actual')),
                    'bar_sfg': safe_float(row.get('Bar SFG')),
                    'qty_allowance': safe_float(row.get('Qty Allowance')),
                    'variant_prod_sfg_pct': safe_float(row.get('Variant Prod SFG (%)')),
                    'variant_fg_pc': safe_float(row.get('Variant FG (PC)')),
                    'variant_fg_pct': safe_float(row.get('Variant FG (%)')),
                    'loss_kg': safe_float(row.get('Lossess FG Result (Kg)')),
                    'loss_pct': safe_float(row.get('Lossess FG Result (%)')),
                    'pc_to_kg_actual': safe_float(row.get('PC to KG (Actual)')),
                    'system_status': safe_str(row.get('System Status')),
                    'ud_status': safe_str(row.get('UD Status')),
                    'pd_manager': safe_str(row.get('PD Manager')),
                    'pd_leader': safe_str(row.get('PD Leader')),
                    'posting_date': reference_date,
                    'source_file': str(file_path.name),
                    'source_row': idx + 2,
                    'raw_data': raw_data,
                }
                raw_records.append(RawZrpp062(**raw_record_data))
                
                # === Execute UPSERT for fact table ===
                fact_params = {
                    'process_order_id': process_order,
                    'batch_id': batch,
                    'material_code': material,
                    'material_description': material_description,
                    'parent_order_id': order_sfg_liquid,
                    'mrp_controller': safe_str(row.get('MRP Controller')),
                    'product_group_1': safe_str(row.get('Product Group 1')),
                    'product_group_2': safe_str(row.get('Product Group 2')),
                    'output_actual_kg': safe_float(row.get('Tonase Alkana(0201)')),
                    'input_actual_kg': safe_float(row.get('GI SFG Liquid to Order')),
                    'process_order_qty': safe_float(row.get('Process Order Qty')),
                    'loss_kg': safe_float(row.get('Lossess FG Result (Kg)')),
                    'loss_pct': safe_float(row.get('Lossess FG Result (%)')),
                    'sg_theoretical': safe_float(row.get('SG Theoretical')),
                    'sg_actual': safe_float(row.get('SG Actual')),
                    'variant_prod_sfg_pct': safe_float(row.get('Variant Prod SFG (%)')),
                    'variant_fg_pct': safe_float(row.get('Variant FG (%)')),
                    'reference_date': reference_date,
                }
                
                self.db.execute(upsert_sql, fact_params)
                self.loaded_count += 1
                
            except Exception as e:
                self.error_count += 1
                self.errors.append(f"Row {idx}: {str(e)}")
        
        # Bulk insert raw records
        if raw_records:
            self.db.bulk_save_objects(raw_records)
        
        self.db.commit()
        
        print(f"  ✓ Upserted {self.loaded_count} rows, Skipped {self.skipped_count}, Errors {self.error_count}")
        return self.get_stats()
    
    def load(self, file_path: Optional[Path] = None) -> Dict[str, int]:
        """
        Legacy load method (V2 compatible) - uses current date as reference_date.
        For V3, prefer load_with_period() method.
        """
        file_path = file_path or self.file_path or EXCEL_FILES.get('zrpp062')
        return self.load_with_period(file_path, date.today())


# Registry of all loaders
LOADERS = {
    'cooispi': CooispiLoader,
    'mb51': Mb51Loader,
    'zrmm024': Zrmm024Loader,
    'zrsd002': Zrsd002Loader,
    'zrsd004': Zrsd004Loader,
    'zrsd006': Zrsd006Loader,
    'zrfi005': Zrfi005Loader,
    'zrpp062': Zrpp062Loader,
    'target': TargetLoader,
}


def get_loader_for_type(file_type: str, file_path: Path, db: Session, mode: str = 'insert', **kwargs):
    """
    Factory function to get appropriate loader for file type
    
    Args:
        file_type: Type of SAP report (cooispi, mb51, zrmm024, etc.)
        file_path: Path to Excel file
        db: Database session
        mode: Load mode ('insert' or 'upsert')
        **kwargs: Additional args passed to loader (e.g., snapshot_date for ZRFI005)
    
    Returns:
        Loader instance
    
    Raises:
        ValueError: If file type is unknown
    
    Example:
        loader = get_loader_for_type('zrfi005', Path('data.xlsx'), db, mode='upsert', snapshot_date=date.today())
    """
    file_type_lower = file_type.lower()
    loader_class = LOADERS.get(file_type_lower)
    
    if not loader_class:
        raise ValueError(f"Unknown file type: {file_type}. Supported: {', '.join(LOADERS.keys())}")
    
    return loader_class(db, mode=mode, file_path=file_path, **kwargs)


def load_all_raw_data(db: Session) -> Dict[str, Dict]:
    """
    Load ALL Excel files into Raw Data Lake
    
    CRITICAL: No filtering - ALL rows, ALL columns
    """
    print("=" * 60)
    print("LOADING RAW DATA (ALL rows, ALL columns)")
    print("=" * 60)
    
    results = {}
    total_loaded = 0
    total_errors = 0
    
    for name, loader_class in LOADERS.items():
        try:
            loader = loader_class(db)
            # CRITICAL: Truncate raw table before loading to prevent duplication
            if hasattr(loader, 'truncate'):
                loader.truncate()
                
            stats = loader.load()
            results[name] = stats
            total_loaded += stats['loaded']
            total_errors += stats['errors']
        except Exception as e:
            print(f"  ✗ Failed to load {name}: {e}")
            results[name] = {'loaded': 0, 'errors': 1, 'error': str(e)}
            total_errors += 1

    
    print("=" * 60)
    print(f"TOTAL: {total_loaded} rows loaded, {total_errors} errors")
    print("=" * 60)
    
    return results
