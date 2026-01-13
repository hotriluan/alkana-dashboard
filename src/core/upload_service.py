"""
Upload Service - File processing and validation

Handles file upload logic:
- File type detection (auto-detect SAP report type)
- File validation (structure, headers, size)
- Background processing with loaders
- Cleanup scheduler
- Transform to fact tables for dashboard

Skills: backend-development, database-operations
"""
import hashlib
import shutil
from pathlib import Path
from datetime import datetime, date, timedelta
from typing import Dict, Optional
import aiofiles
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from src.db.models import UploadHistory
from src.etl.loaders import get_loader_for_type, Zrfi005Loader
from src.etl.transform import Transformer


def compute_file_hash(file_path: Path) -> str:
    """Compute MD5 hash of file for duplicate detection"""
    md5_hash = hashlib.md5()
    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(4096), b""):
            md5_hash.update(chunk)
    return md5_hash.hexdigest()


def detect_file_type(file_path: Path) -> str:
    """
    Auto-detect SAP report type by analyzing Excel headers
    
    Returns:
        File type: COOISPI, MB51, ZRMM024, ZRSD002, ZRSD004, ZRSD006, ZRFI005, TARGET
    
    Raises:
        ValueError: If file type cannot be determined
    """
    try:
        # Don't use read_only=True - it doesn't read headers correctly for some files
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        # Read first row headers (up to 30 columns)
        headers = []
        for col in range(1, min(ws.max_column or 30, 30) + 1):
            cell_value = ws.cell(1, col).value
            if cell_value:
                headers.append(str(cell_value).lower().strip())
        
        wb.close()
        
        headers_str = '|'.join(headers)
        
        # Match patterns based on characteristic columns
        if 'order' in headers_str and 'batch' in headers_str and 'material number' in headers_str:
            return 'COOISPI'
        elif 'material document' in headers_str and 'material' in headers_str:
            return 'MB51'
        elif 'purch. order' in headers_str or 'purch order' in headers_str:
            return 'ZRMM024'
        elif 'billing doc' in headers_str or 'billing document' in headers_str:
            return 'ZRSD002'
        elif 'delivery' in headers_str and 'sold-to party' in headers_str:
            return 'ZRSD004'
        elif ('material' in headers_str or 'material code' in headers_str) and 'distribution channel' in headers_str and ('ph 1' in headers_str or 'ph 2' in headers_str or 'ph1' in headers_str or 'ph2' in headers_str):
            return 'ZRSD006'
        elif 'customer name' in headers_str and 'total target' in headers_str and 'total realization' in headers_str:
            return 'ZRFI005'
        elif 'salesman name' in headers_str and 'semester' in headers_str and 'target' in headers_str:
            return 'TARGET'
        else:
            raise ValueError(f"Unknown file type. Headers found: {', '.join(headers[:10])}")
    
    except Exception as e:
        raise ValueError(f"Failed to detect file type: {str(e)}")


def validate_file_structure(file_path: Path, expected_type: Optional[str] = None) -> Dict:
    """
    Validate Excel file structure
    
    Returns:
        dict: {valid: bool, file_type: str, rows: int, columns: int, error: str}
    """
    try:
        # Detect file type
        file_type = detect_file_type(file_path)
        
        # Validate against expected type if provided
        if expected_type and expected_type.upper() != file_type:
            return {
                'valid': False,
                'file_type': file_type,
                'error': f"Expected {expected_type} but detected {file_type}"
            }
        
        # Get row/column count
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.max_row or 0
        cols = ws.max_column or 0
        wb.close()
        
        # Validate minimum requirements
        if rows < 2:
            return {
                'valid': False,
                'file_type': file_type,
                'error': f"File has only {rows} rows (needs at least 2: header + data)"
            }
        
        if cols < 3:
            return {
                'valid': False,
                'file_type': file_type,
                'error': f"File has only {cols} columns (needs at least 3)"
            }
        
        return {
            'valid': True,
            'file_type': file_type,
            'rows': rows,
            'columns': cols,
            'error': None
        }
    
    except Exception as e:
        return {
            'valid': False,
            'file_type': None,
            'error': str(e)
        }


async def save_upload_file(upload_file, destination: Path) -> int:
    """
    Save uploaded file to destination
    
    Returns:
        File size in bytes
    """
    destination.parent.mkdir(parents=True, exist_ok=True)
    
    async with aiofiles.open(destination, 'wb') as out_file:
        content = await upload_file.read()
        await out_file.write(content)
    
    return len(content)


async def process_file(upload_id: int, file_path: Path, db: Session) -> Dict:
    """
    Process uploaded file with appropriate loader
    
    This function:
    1. Validates file structure
    2. Detects file type
    3. Calls appropriate loader with upsert mode
    4. Updates upload_history with results
    
    Returns:
        Processing statistics
    """
    upload = db.query(UploadHistory).filter_by(id=upload_id).first()
    if not upload:
        raise ValueError(f"Upload {upload_id} not found")
    
    try:
        # Update status to processing
        upload.status = 'processing'
        db.commit()
        
        # Validate file
        validation = validate_file_structure(file_path)
        if not validation['valid']:
            raise ValueError(validation['error'])
        
        file_type = validation['file_type']
        upload.file_type = file_type
        
        # Get snapshot_date from upload record (if provided by user)
        snapshot_date = upload.snapshot_date or date.today()
        
        # Use upsert mode for all loaders (UPDATE existing, INSERT new, SKIP unchanged)
        mode = 'upsert'
        db.commit()
        
        # Get appropriate loader (REUSE existing loaders)
        if file_type == 'ZRFI005':
            # AR special handling: pass snapshot_date to loader
            loader = Zrfi005Loader(db, mode=mode, file_path=file_path)
            stats = loader.load(snapshot_date=snapshot_date)
        else:
            # Standard loaders with upsert mode
            loader = get_loader_for_type(file_type.lower(), file_path, db, mode=mode)
            stats = loader.load()
        
        # Transform raw data to fact tables for dashboard
        print(f"  🔄 Transforming {file_type} to fact tables...")
        transformer = Transformer(db)
        
        # Transform only the relevant table based on file type
        if file_type == 'COOISPI':
            transformer.transform_cooispi()
            # COOISPI impacts production chains, lead time, and alerts
            transformer.build_production_chains()
            transformer.calculate_p02_p01_yields()
            transformer.transform_lead_time()  # Calculate transit days for P01 batches
            transformer.detect_alerts()
        elif file_type == 'MB51':
            transformer.transform_mb51()
            # MB51 impacts production chains, lead time, and alerts
            transformer.build_production_chains()
            transformer.calculate_p02_p01_yields()
            transformer.transform_lead_time()
            transformer.detect_alerts()
        elif file_type == 'ZRMM024':
            transformer.transform_zrmm024()
            # ZRMM024 impacts lead time (purchase time)
            transformer.transform_lead_time()
        elif file_type == 'ZRSD002':
            transformer.transform_zrsd002()
            transformer.build_uom_conversion()  # Update UOM conversion from billing data
            # ZRSD002 impacts lead time (sales order data)
            transformer.transform_lead_time()
        elif file_type == 'ZRSD004':
            transformer.transform_zrsd004()
        elif file_type == 'ZRSD006':
            # ZRSD006 is used by lead time for channel lookup
            transformer.transform_lead_time()
        elif file_type == 'ZRFI005':
            # Pass snapshot_date to transformer to ensure correct data is aggregated
            transformer.transform_zrfi005(target_date=snapshot_date.isoformat() if snapshot_date else None)
        elif file_type == 'TARGET':
            transformer.transform_target()
        
        print(f"  ✓ Transform completed")
        
        # Update statistics
        upload.status = 'completed'
        upload.rows_loaded = stats.get('loaded', 0)
        upload.rows_updated = stats.get('updated', 0)
        upload.rows_skipped = stats.get('skipped', 0)
        # errors is returned as a list; store the count in the integer column
        error_list = stats.get('errors', []) or []
        upload.rows_failed = len(error_list)
        upload.processed_at = datetime.utcnow()
        db.commit()
        
        return stats
    
    except Exception as e:
        # Rollback on error
        db.rollback()
        upload.status = 'failed'
        upload.error_message = str(e)
        upload.processed_at = datetime.utcnow()
        db.commit()
        raise


def cleanup_old_uploads(uploads_dir: Path, max_age_hours: int = 24):
    """
    Delete uploaded files older than max_age_hours
    
    Called periodically to prevent disk space issues
    """
    if not uploads_dir.exists():
        return
    
    cutoff_time = datetime.now() - timedelta(hours=max_age_hours)
    deleted_count = 0
    
    for file_path in uploads_dir.glob("*.xlsx"):
        # Check file modification time
        file_mtime = datetime.fromtimestamp(file_path.stat().st_mtime)
        if file_mtime < cutoff_time:
            try:
                file_path.unlink()
                deleted_count += 1
            except Exception as e:
                print(f"Failed to delete {file_path}: {e}")
    
    return deleted_count
