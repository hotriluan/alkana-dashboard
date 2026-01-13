#!/usr/bin/env python3
"""Check if ZRSD006 data already exists in database"""
import sys
from pathlib import Path
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))

from src.db.connection import SessionLocal
from src.db.models import RawZrsd006

db = SessionLocal()

try:
    # Count total records
    total = db.query(RawZrsd006).count()
    print(f"Total RawZrsd006 records in database: {total}")
    
    if total == 0:
        print("✓ Database is EMPTY → Upload should load all rows")
    else:
        print(f"⚠️ Database has {total} records\n")
        
        # Show sample records
        print("Sample records in database (first 3):")
        for rec in db.query(RawZrsd006).limit(3).all():
            print(f"  Material: {rec.material}, Dist Channel: {rec.dist_channel}")
        
        # Check Excel files
        print("\n\n=== Excel File Stats ===")
        excel_files = ['demodata/11.XLSX', 'demodata/12.XLSX', 'demodata/13.XLSX', 'demodata/15.XLSX']
        total_rows = 0
        
        for file in excel_files:
            p = Path(file)
            if p.exists():
                wb = load_workbook(p, data_only=True)
                ws = wb.active
                row_count = ws.max_row - 1  # Minus header
                total_rows += row_count
                print(f"{p.name}: {row_count} data rows")
                wb.close()
        
        print(f"\nTotal rows in all 4 Excel files: {total_rows}")
        
        if total == total_rows:
            print("\n✓ Database matches Excel → Skip is CORRECT (all duplicates)")
        else:
            print(f"\n⚠️ Mismatch: {total} in DB vs {total_rows} in Excel")
            print("   → Check for partial loads or filtering")

finally:
    db.close()
