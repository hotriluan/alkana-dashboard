import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

file_path = Path('demodata/zrsd004.XLSX')

print("=" * 80)
print("🔍 ANALYZING EXCEL FILE STRUCTURE")
print("=" * 80)

# Method 1: Read with pandas (default header=0)
print("\n📊 Method 1: Pandas read_excel (header=0)")
df1 = pd.read_excel(file_path, header=0, nrows=5, dtype=str)
print(f"Columns: {list(df1.columns)[:5]}...")
print(f"First row data:")
print(df1.head(1))

# Method 2: Read with pandas (header=None to see raw structure)
print("\n📊 Method 2: Pandas read_excel (header=None, first 5 rows)")
df2 = pd.read_excel(file_path, header=None, nrows=5, dtype=str)
print(df2)

# Method 3: Openpyxl to see actual Excel structure
print("\n📊 Method 3: Openpyxl (first 5 rows)")
wb = load_workbook(file_path, data_only=True)
ws = wb.active
for row_idx in range(1, 6):
    row_values = [ws.cell(row_idx, col).value for col in range(1, 10)]  # First 9 columns
    print(f"Row {row_idx}: {row_values}")
wb.close()

# Check specific delivery records
print("\n" + "=" * 80)
print("🔍 CHECKING SPECIFIC DELIVERIES")
print("=" * 80)

df_full = pd.read_excel(file_path, header=0, dtype=str)
print(f"\nTotal records: {len(df_full)}")

for delivery_num in ['1910053734', '1910053733', '1910053732']:
    print(f"\n🔎 Delivery: {delivery_num}")
    matches = df_full[df_full['Delivery'].astype(str).str.strip() == delivery_num]
    if len(matches) > 0:
        print(f"  Found {len(matches)} records")
        for idx, row in matches.iterrows():
            print(f"  Row {idx + 2}: Delivery Date = '{row['Delivery Date']}', Line Item = {row['Line Item']}")
    else:
        print(f"  ❌ NOT FOUND")
